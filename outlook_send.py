import win32com.client as win32
import os
from openpyxl import load_workbook
import time
import random


def findFile():
    while True:
        try:
            m_path = os.getcwd()
            m_path = m_path.split('\\')[1:3]
            users = m_path[0]
            user = m_path[1]
            m_table = f"C:\{users}" + f"\{user}\Gamma International Group\ONE STORY - Документы\Change\Automation\Разработки\Рассылка\Письма_стандартизация.xlsx"
            m_pdf = f"C:\{users}" + f"\{user}\Gamma International Group\ONE STORY - Документы\Change\Automation\Разработки\Рассылка\Решения для бизнеса - Skolkovo NextGen.pdf"
            return m_table, m_pdf
        except IndexError:
            continue


def openTable(table):
    wb = load_workbook(table)
    sheet = wb['1']
    maxRow = sheet.max_row
    return maxRow, sheet


def waitSec():
    sec = random.random()
    time.sleep(sec)


def openData(sheet, i):
    email = sheet['A' + str(i)].value
    return email


def dropEmails(email):
    emails = []
    email = email.replace(' ', '')
    while ';' in email:
        start = 0
        end = email.find(';')
        emails.append(email[start: end])
        start = end + 1
        email = email[start: len(email)]
    emails.append(email)
    return emails


def sendMessages(emails, pdf):
    outlook_app = win32.Dispatch('Outlook.Application')

    send_account = None
    for account in outlook_app.Session.Accounts:
        if account.DisplayName == 'islamova@onestory.pro':
            send_account = account
            break

    for i in range(len(emails)):
        waitSec()
        mail_item = outlook_app.CreateItem(0)  # 0: olMailItem
        mail_item._oleobj_.Invoke(*(64209, 0, 8, 0, send_account))
        mail_item.Recipients.Add(emails[i])
        print("Сообщение будет отправлено на почту - " + emails[i])
        mail_item.Subject = 'Предложение о сотрудничестве - Skolkovo NextGen'
        mail_item.BodyFormat = 2
        mail_item.HTMLBody = '''...''' # 2: Html-код письма
        mail_item.Attachments.Add(Source=pdf)
        mail_item.Send()
        print("Сообщение отправлено!")


def main():
    table, pdf = findFile()
    maxRow, sheet = openTable(table)
    for i in range(1, maxRow + 1):
        email = openData(sheet, i)
        emails = dropEmails(email)
        sendMessages(emails, pdf)


if __name__ == "__main__":
    try:
        main()
    except:
        input()